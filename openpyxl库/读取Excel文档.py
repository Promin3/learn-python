import openpyxl as pxl

book = pxl.load_workbook(
    r"C:\Users\朱俊杰\Desktop\2022华数杯\2022年“华数杯”全国大学生数学建模竞赛赛题\2022年“华数杯”全国大学生数学建模竞赛赛题\2022年“华数杯”全国大学生数学建模竞赛C题\C题数据.xlsx")
sheet0 = book.worksheets[0]
print(sheet0.title)
print(sheet0.min_row, sheet0.max_row)
print(sheet0.min_column, sheet0.max_column)
for row in sheet0.rows:
    for cell in row:
        print(cell.value, " ", end="")
    print("")
for cell in sheet0["G"]:
    print(cell.value)
for cell in sheet0[3]:
    print(cell.value, " ", end="")
for cell in sheet0[3]:
    print(cell.value, type(cell.value), cell.coordinate, cell.col_idx, cell.number_format)
